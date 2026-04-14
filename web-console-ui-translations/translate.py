#!/usr/bin/env python3
"""
Avatour Web Console UI Translation Script
==========================================
Translates en.json to Italian and German using the DeepL API.

Approach:
  1. Mask all placeholders ($t(), {{var}}, <0>, SuperFreeze etc.) before
     sending to DeepL so they are never translated or reordered.
  2. Translate via DeepL with formal register (Sie / Lei).
  3. Apply post-processing fixes for known bad translations — this is more
     reliable than the DeepL glossary API which requires a paid plan to
     work consistently.
  4. Output it.json, de.json, and translations.xlsx for review.

Usage:
    python translate.py

Environment variables:
    DEEPL_API_KEY   DeepL API key (set as GitHub Actions secret)
"""

import json
import os
import re
import sys
import time
from pathlib import Path

# -- Dependencies --------------------------------------------------------------
try:
    import requests
except ImportError:
    os.system(f"{sys.executable} -m pip install requests --quiet")
    import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

# -- Configuration -------------------------------------------------------------

DEEPL_API_KEY      = os.environ.get("DEEPL_API_KEY", "")
DEEPL_API_URL      = "https://api.deepl.com/v2/translate"
DEEPL_GLOSSARY_URL = "https://api.deepl.com/v2/glossaries"

SOURCE_FILE    = Path(__file__).parent / "en.json"
PREV_FILE      = Path(__file__).parent / "dist" / "en-prev.json"
DIST_DIR       = Path(__file__).parent / "dist"

LANGUAGES = {
    "IT": "Italian",
    "DE": "German",
}

# -- Pre-translations ----------------------------------------------------------
# These keys are set BEFORE DeepL runs, bypassing translation entirely.
# Used for single words that DeepL consistently mistranslates.

PRE_TRANSLATIONS = {
    "DE": {
        "volume":    "Lautstärke",
        "assets":    "Assets",
        "asset":     "Asset",
        "history":   "Verlauf",
        "exit":      "Beenden",
        "play":      "Wiedergabe",
        "cancel":    "Abbrechen",
        "analytics": "Analysen",
        "filters":   "Filter",
        "host":      "Host",
        "hosts":     "Hosts",
        "meeting":   "Meeting",
        "session":   "Meeting",
        "change":    "Ändern",
        "deactivate": "Deaktivieren",
        "reactivate": "Reaktivieren",
        "new_meeting": "neues Meeting",
    },
    "IT": {
        "assets":      "asset",
        "asset":       "asset",
        "history":     "cronologia",
        "host":        "host",
        "hosts":       "host",
        "exit":        "Esci",
        "play":        "Riproduci",
        "cancel":      "Annulla",
        "change":      "Cambia",
        "analytics":   "analisi",
        "filters":     "filtri",
        "superfreeze": "SuperFreeze",
        "submit":      "Invia",
    },
}

# -- Glossaries (DeepL Pro) ----------------------------------------------------
# Used to enforce correct translations at the DeepL API level.
# With Pro, glossaries work reliably via the API.

GLOSSARIES = {
    "DE": {
        # Meeting pluralisation — DeepL incorrectly uses "Meetingen"
        "meetings":  "Meetings",
        "meeting":   "Meeting",
        # Ensure consistent capitalisation
        "Meeting":   "Meeting",
        "Meetings":  "Meetings",
    },
    "IT": {
        # No glossary entries needed currently
    },
}

# -- Post-processing fixes -----------------------------------------------------
# Applied after DeepL translation to correct known bad translations.
# Each entry is (wrong_value, correct_value) — exact string match on the
# full translated value only (not substring replacement within longer strings).
#
# To add a new fix: add a tuple ("wrong DeepL output", "correct translation")

FIXES = {
    "DE": [
        # Single-word / short UI labels — exact value fixes
        ("band",             "Lautstärke"),       # volume
        ("vermögenswerte",   "Assets"),            # assets
        ("vermögen",         "Asset"),             # asset
        ("geschichte",       "Verlauf"),           # history
        ("ausgang",          "Beenden"),           # exit
        ("spielen",          "Wiedergabe"),        # play
        ("abbrechen",        "Abbrechen"),         # cancel (already correct casing)
        ("analytik",         "Analysen"),          # analytics
        ("filtert",          "Filter"),            # filters
        ("gastgeber",        "Host"),              # host
        ("treffen",          "Meeting"),           # meeting/session
        ("sitzung",          "Meeting"),           # meeting (alt)
        # Compound fixes — verb forms on buttons
        ("Ändern Sie",       "Ändern"),            # change (verb)
        ("deaktivieren Sie", "Deaktivieren"),      # deactivate
        ("reaktivieren Sie", "Reaktivieren"),      # reactivate
        # Broken asset-related strings
        ("ANZAHL DER VERMÖGENSWERTE", "ANZAHL DER ASSETS"),
        ("lebenslauf ein Vorteil",    "Asset fortsetzen"),
        ("einen Vermögenswert angehalten", "Asset pausiert"),
        ("Möchten Sie das SuperFreeze als Anlage behalten?",
         "Möchten Sie den SuperFreeze als Asset behalten?"),
        # Waiting for guide (Reiseführer = travel guide book)
        ("Warten auf den Reiseführer...", "Warten auf den Host..."),
        # asset singular still being retranslated
        ("Vermögenswert", "Asset"),
        # new_meeting grammar
        ("neue Meeting", "neues Meeting"),
    ],
    "IT": [
        # Single-word / short UI labels — exact value fixes
        ("patrimonio",   "asset"),          # assets
        ("attività",     "asset"),          # asset
        ("storia",       "cronologia"),     # history
        ("padroni di casa", "host"),        # hosts
        ("uscita",       "Esci"),           # exit
        ("giocare",      "Riproduci"),      # play
        ("annullamento", "Annulla"),        # cancel (noun→verb)
        ("presentare",   "Presente"),       # present (verb)
        ("Tutti gli ospiti", "Tutti gli host"),  # all_hosts
        # Compound fixes
        ("Cambiamento",  "Cambia"),         # change (noun→verb)
        ("superGelo",    "SuperFreeze"),    # product name
        ("SuperCongelamento accettato", "SuperFreeze accettato"),
        ("Richiesta di supercongelamento rifiutata",
         "Richiesta SuperFreeze rifiutata"),
        # SuperFreeze gender fix
        ("la SuperFreeze", "il SuperFreeze"),
        ("La SuperFreeze", "Il SuperFreeze"),
        # waiting_for_guide
        ("In attesa della guida...", "In attesa dell'host..."),
        # Asset-related strings
        ("NUMERO DI BENI",   "NUMERO DI ASSET"),
        ("Numero di Attività per tipo", "Numero di asset per tipo"),
        ("il curriculum è una risorsa", "riprende un asset"),
        ("Vuole mantenere il SuperFreeze come risorsa?",
         "Vuole mantenere il SuperFreeze come asset?"),
        ("Desidera mantenere la SuperFreeze come risorsa?",
         "Vuole mantenere il SuperFreeze come asset?"),
        ("Tutte le note associate a questo bene saranno cancellate quando il layout viene modificato",
         "Tutte le note associate a questo asset saranno cancellate quando il layout viene modificato"),
        # Stray @ fixes — targeted replacements for known affected strings
        ("In attesa che Host inizi la riunione...",
         "In attesa che l'host inizi la riunione..."),
        ("SuperFreeze ha richiesto", "SuperFreeze richiesto"),
    ],
}

# Substring fixes — applied within longer string values (not exact match)
# Format: (substring_to_find, replacement) per language
SUBSTRING_FIXES = {
    "DE": [
        ("Gastgeber",    "Host"),          # host within sentences
        ("Treffen",      "Meeting"),       # meeting within sentences
        ("Sitzung",      "Meeting"),       # meeting (alt) within sentences
        ("Vermögenswert", "Asset"),        # asset within sentences
    ],
    "IT": [
        # Note: stray @ removal is handled in apply_fixes() using regex
        # to avoid stripping @ from email addresses like support@avatour.live
    ],
}


def apply_fixes(translated_dict, lang_code):
    """
    Apply post-processing fixes to translated values.
    1. Exact-match fixes (full value match, case-insensitive)
    2. Substring fixes
    3. Language-specific regex fixes
    """
    exact_fixes = FIXES.get(lang_code, [])
    substr_fixes = SUBSTRING_FIXES.get(lang_code, [])
    fixed_count = 0
    result = {}

    for key, value in translated_dict.items():
        new_value = value

        # Exact match fixes (case-insensitive comparison)
        for wrong, correct in exact_fixes:
            if new_value.lower() == wrong.lower():
                new_value = correct
                fixed_count += 1
                break

        # Substring fixes
        if new_value == value:
            for wrong, correct in substr_fixes:
                if wrong in new_value:
                    new_value = new_value.replace(wrong, correct)
                    fixed_count += 1

        # Post-restore fixes for all languages
        # DeepL strips @ from email addresses even inside masked tags — restore them
        if "supportavatour.live" in new_value:
            new_value = new_value.replace("supportavatour.live", "support@avatour.live")
            fixed_count += 1

        # Italian-specific fixes
        if lang_code == "IT":
            # Remove stray @ DeepL inserts — appears after restored placeholders
            # e.g. "$t(avatour)@ account" or "</3>@ di" or "</1>@\n"
            cleaned = re.sub(r'(\$t\([^)]+\)|</?\d+>|</?\w+>)@', r'\1', new_value)
            if cleaned != new_value:
                new_value = cleaned
                fixed_count += 1
            # Also catch trailing @ at end of string
            if new_value.endswith("@"):
                new_value = new_value.rstrip("@").rstrip()
                fixed_count += 1
            # Sentence-level fixes
            it_sentence_fixes = [
                ("In attesa che Host inizi", "In attesa che l'host inizi"),
                ("SuperFreeze ha richiesto", "SuperFreeze richiesto"),
            ]
            for wrong, correct in it_sentence_fixes:
                if wrong in new_value:
                    new_value = new_value.replace(wrong, correct)
                    fixed_count += 1

        # Capitalise first letter for DE and IT — German nouns require capitals,
        # and Italian labels look more consistent capitalised.
        # Only applies when the first character is a lowercase letter.
        # Skips $t() references, strings starting with special chars, etc.
        if lang_code in ("DE", "IT"):
            if new_value and new_value[0].islower():
                new_value = new_value[0].upper() + new_value[1:]
                fixed_count += 1

        result[key] = new_value

    print(f"  Post-processing: {fixed_count} fixes applied.")
    return result


# -- Skip pattern --------------------------------------------------------------
SKIP_TRANSLATION_PATTERN = re.compile(
    r"^\$t\([^)]+\)$"
    r"|^\$t\([^)]+\)\s+\$t\([^)]+\)$"
)

# -- Placeholder protection ----------------------------------------------------
PLACEHOLDER_PATTERNS = [
    re.compile(r"\$t\([^)]+\)"),
    re.compile(r"\{\{[^}]+\}\}"),
    re.compile(r"<\d+>[^<]*</\d+>"),
    re.compile(r"<\d+></\d+>"),
    re.compile(r"<\d+>"),
    re.compile(r"</\d+>"),
    re.compile(r"<[a-zA-Z]\w*>[^<]*</[a-zA-Z]\w*>"),
    re.compile(r"<[a-zA-Z]\w*>"),
    re.compile(r"</[a-zA-Z]\w*>"),
    re.compile(r"support@avatour\.live"),
    re.compile(r"SuperFreeze"),
]


def mask_placeholders(text):
    matches = []
    for pattern in PLACEHOLDER_PATTERNS:
        for m in pattern.finditer(text):
            matches.append((m.start(), m.end(), m.group()))

    matches.sort(key=lambda x: x[0])
    non_overlapping = []
    last_end = -1
    for start, end, value in matches:
        if start >= last_end:
            non_overlapping.append((start, end, value))
            last_end = end

    placeholders = [v for _, _, v in non_overlapping]
    masked = text
    for i, (start, end, _) in enumerate(reversed(non_overlapping)):
        idx = len(non_overlapping) - 1 - i
        masked = masked[:start] + f"@@PH{idx}@@" + masked[end:]

    return masked, placeholders


def restore_placeholders(text, placeholders):
    result = text
    for i, value in enumerate(placeholders):
        result = result.replace(f"@@PH{i}@@", value)
    return result


# -- DeepL Glossary API (Pro) --------------------------------------------------

def ensure_glossary(lang_code):
    """
    Create or retrieve a DeepL glossary for lang_code.
    Returns glossary_id or None if no entries defined for this language.
    """
    entries = GLOSSARIES.get(lang_code, {})
    if not entries:
        return None

    if not DEEPL_API_KEY:
        return None

    headers = {
        "Authorization": f"DeepL-Auth-Key {DEEPL_API_KEY}",
        "Content-Type": "application/json",
    }

    # Delete any existing Avatour UI glossary for this language
    resp = requests.get(DEEPL_GLOSSARY_URL, headers=headers, timeout=15)
    if resp.status_code == 200:
        for g in resp.json().get("glossaries", []):
            if (g.get("name", "").startswith("avatour-ui-")
                    and g.get("target_lang", "").upper() == lang_code.upper()):
                requests.delete(
                    f"{DEEPL_GLOSSARY_URL}/{g['glossary_id']}",
                    headers=headers, timeout=15
                )

    # Create glossary using form-encoded data (DeepL v2 requirement)
    tsv = "\n".join(f"{k}\t{v}" for k, v in entries.items())
    resp = requests.post(
        DEEPL_GLOSSARY_URL,
        headers={"Authorization": f"DeepL-Auth-Key {DEEPL_API_KEY}"},
        data={
            "name":           f"avatour-ui-{lang_code.lower()}",
            "source_lang":    "EN",
            "target_lang":    lang_code,
            "entries":        tsv,
            "entries_format": "tsv",
        },
        timeout=15
    )

    if resp.status_code not in (200, 201):
        print(f"  Warning: glossary creation failed ({resp.status_code}) — continuing without glossary")
        return None

    glossary_id = resp.json().get("glossary_id")
    print(f"  Glossary created for {lang_code}: {len(entries)} terms (id={glossary_id})")

    # Wait for glossary to be ready
    for _ in range(10):
        time.sleep(1)
        check = requests.get(
            f"{DEEPL_GLOSSARY_URL}/{glossary_id}",
            headers={"Authorization": f"DeepL-Auth-Key {DEEPL_API_KEY}"},
            timeout=15
        )
        if check.status_code == 200 and check.json().get("ready"):
            break

    return glossary_id


def delete_glossary(glossary_id):
    """Remove a glossary from DeepL after use."""
    if not glossary_id:
        return
    requests.delete(
        f"{DEEPL_GLOSSARY_URL}/{glossary_id}",
        headers={"Authorization": f"DeepL-Auth-Key {DEEPL_API_KEY}"},
        timeout=15
    )


# -- DeepL Translation API -----------------------------------------------------

def translate_batch(texts, target_lang, glossary_id=None):
    """Translate a list of strings via DeepL with formal register."""
    if not DEEPL_API_KEY:
        raise ValueError("DEEPL_API_KEY environment variable is not set.")

    headers = {
        "Authorization": f"DeepL-Auth-Key {DEEPL_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "text":                texts,
        "source_lang":         "EN",
        "target_lang":         target_lang,
        "formality":           "prefer_more",
        "preserve_formatting": True,
    }
    if glossary_id:
        payload["glossary_id"] = glossary_id

    response = requests.post(DEEPL_API_URL, headers=headers,
                             json=payload, timeout=30)
    if response.status_code != 200:
        raise RuntimeError(
            f"DeepL API error {response.status_code}: {response.text}"
        )
    return [t["text"] for t in response.json()["translations"]]


def translate_all(strings_dict, target_lang, glossary_id=None):
    """Translate all values, skipping pure $t() refs and protecting placeholders."""
    # Apply pre-translations — these keys bypass DeepL entirely
    pre = PRE_TRANSLATIONS.get(target_lang, {})
    strings_dict = {k: pre[k] if k in pre else v
                    for k, v in strings_dict.items()}
    print(f"  Pre-translated {len(pre)} keys, bypassing DeepL.")

    keys   = list(strings_dict.keys())
    values = list(strings_dict.values())

    to_translate_indices = []
    to_translate_masked  = []
    placeholders_map     = {}

    for i, value in enumerate(values):
        if SKIP_TRANSLATION_PATTERN.match(value):
            continue
        if keys[i] in pre:  # pre-translated — skip DeepL entirely
            continue
        masked, placeholders = mask_placeholders(value)
        to_translate_indices.append(i)
        to_translate_masked.append(masked)
        placeholders_map[i] = placeholders

    print(f"  Translating {len(to_translate_indices)} strings "
          f"({len(values) - len(to_translate_indices)} skipped as pure references)...")

    BATCH_SIZE = 50
    translated_masked = []

    for batch_start in range(0, len(to_translate_masked), BATCH_SIZE):
        batch         = to_translate_masked[batch_start:batch_start + BATCH_SIZE]
        batch_num     = batch_start // BATCH_SIZE + 1
        total_batches = (len(to_translate_masked) + BATCH_SIZE - 1) // BATCH_SIZE
        print(f"    Batch {batch_num}/{total_batches} ({len(batch)} strings)...")
        translated_batch = translate_batch(batch, target_lang, glossary_id)
        translated_masked.extend(translated_batch)
        if batch_start + BATCH_SIZE < len(to_translate_masked):
            time.sleep(0.2)

    result = list(values)
    for j, i in enumerate(to_translate_indices):
        restored = restore_placeholders(translated_masked[j], placeholders_map[i])
        result[i] = restored

    return dict(zip(keys, result))


# -- Excel output --------------------------------------------------------------

def write_excel(en_dict, translations, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"

    header_fill = PatternFill("solid", fgColor="132A39")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Key", "English"] + [LANGUAGES[l] for l in translations]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    for letter in ["C", "D", "E"]:
        ws.column_dimensions[letter].width = 50

    alt_fill = PatternFill("solid", fgColor="F7F8F9")

    for row_idx, key in enumerate(en_dict.keys(), 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        ws.cell(row=row_idx, column=1, value=key).fill = fill
        en_cell = ws.cell(row=row_idx, column=2, value=en_dict[key])
        en_cell.fill = fill
        en_cell.alignment = Alignment(wrap_text=True)
        for col_idx, (lang_code, lang_dict) in enumerate(translations.items(), 3):
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=lang_dict.get(key, ""))
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"  Excel review file written: {output_path.name}")


# -- Main ----------------------------------------------------------------------

def load_prev_english():
    """Load en-prev.json if it exists, otherwise return empty dict."""
    if PREV_FILE.exists():
        with open(PREV_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def find_changed_keys(en_dict, en_prev):
    """
    Return the set of keys whose English value has changed since the last run.
    New keys (not in en-prev) are also included.
    If en-prev does not exist, all keys are considered changed (full run).
    """
    if not en_prev:
        return set(en_dict.keys())
    changed = set()
    for key, value in en_dict.items():
        if en_prev.get(key) != value:
            changed.add(key)
    return changed


def main():
    print("Avatour Web Console UI Translation")
    print("====================================")

    if not SOURCE_FILE.exists():
        print(f"ERROR: Source file not found: {SOURCE_FILE}")
        sys.exit(1)

    with open(SOURCE_FILE, encoding="utf-8-sig") as f:
        en_dict = json.load(f)

    print(f"Loaded {len(en_dict)} strings from en.json")
    DIST_DIR.mkdir(exist_ok=True)

    # Delta detection — only translate keys that changed since last run
    en_prev   = load_prev_english()
    changed   = find_changed_keys(en_dict, en_prev)
    is_full   = not en_prev

    if is_full:
        print(f"No en-prev.json found — running full translation ({len(en_dict)} keys).")
    else:
        print(f"Delta detected: {len(changed)} of {len(en_dict)} keys changed since last run.")
        if not changed:
            print("Nothing to translate — en.json unchanged. Exiting.")
            # Still update en-prev.json and exit cleanly
            with open(PREV_FILE, "w", encoding="utf-8") as f:
                json.dump(en_dict, f, ensure_ascii=False, indent=2)
            return

    all_translations = {}

    for lang_code, lang_name in LANGUAGES.items():
        print(f"\nTranslating to {lang_name} ({lang_code})...")

        out_path = DIST_DIR / f"{lang_code.lower()}.json"

        # Load existing translations to preserve manual edits
        if out_path.exists() and not is_full:
            with open(out_path, encoding="utf-8") as f:
                existing = json.load(f)
            print(f"  Loaded {len(existing)} existing translations from {out_path.name}")
        else:
            existing = {}

        # Build a subset dict of only the changed keys to translate
        changed_dict = {k: v for k, v in en_dict.items() if k in changed}

        # Create glossary (Pro feature — enforces correct terminology)
        glossary_id = ensure_glossary(lang_code)

        # Translate only the changed keys
        new_translations = translate_all(changed_dict, lang_code, glossary_id)
        new_translations = apply_fixes(new_translations, lang_code)

        # Clean up glossary
        delete_glossary(glossary_id)

        # Merge: existing translations take the base, changed keys are overwritten
        merged = dict(existing)
        merged.update(new_translations)

        # Preserve key order from en.json
        translated = {k: merged[k] for k in en_dict if k in merged}

        all_translations[lang_code] = translated

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(translated, f, ensure_ascii=False, indent=2)
        print(f"  JSON written: {out_path.name} ({len(new_translations)} keys updated)")

    # Save en-prev.json for next run
    with open(PREV_FILE, "w", encoding="utf-8") as f:
        json.dump(en_dict, f, ensure_ascii=False, indent=2)
    print(f"\n  Snapshot saved: dist/en-prev.json")

    print("\nWriting Excel review file...")
    write_excel(en_dict, all_translations, DIST_DIR / "translations.xlsx")

    print("\nDone!")
    print("  dist/it.json           — Italian translations")
    print("  dist/de.json           — German translations")
    print("  dist/en-prev.json      — English snapshot for delta detection")
    print("  dist/translations.xlsx — Combined review file (EN / IT / DE)")


if __name__ == "__main__":
    main()
